import streamlit as st
import pandas as pd
import openpyxl
import io
import os
from PIL import Image

# 1. Configuração da Página Web
st.set_page_config(
    page_title="Renato Frigotudo & Associados",
    layout="wide",
    initial_sidebar_state="expanded"
)

ARQUIVO_TEMPLATE = "GOIAS NOVO DFC PROJETADO-SEM CORTE.xlsx"

# --- CABEÇALHO PERSONALIZADO ---
col_logo, col_titulo = st.columns([1, 4])

with col_logo:
    try:
        logo = Image.open("logo.JPG")
        st.image(logo, width=150)
    except Exception:
        try:
            logo = Image.open("input_file_1.png")
            st.image(logo, width=150)
        except Exception:
            st.subheader("🥩 LOGO")

with col_titulo:
    st.title("Renato Frigotudo & Associados")
    st.subheader("Simulador Dinâmico de Fluxo de Caixa Projetado")

st.markdown("---")

# 2. BARRA LATERAL (Campos Digitados de Precisão)
st.sidebar.header("⚙️ Configurações da Simulação")

nome_cliente = st.sidebar.text_input("Nome do Cliente", value="RENATO")

arquivo_carregado = st.sidebar.file_uploader(
    "Carrega a Planilha de Dados Históricos (.xlsx)", 
    type=["xlsx"]
)

st.sidebar.markdown("### 📊 Digite os percentuais de simulação:")

# Campos de digitação manual direta para precisão
perc_receita = st.sidebar.number_input(
    "Aumento/Diminuição da Receita (%)", 
    min_value=-100.0, 
    max_value=300.0, 
    value=3.0, 
    step=0.1,
    format="%.2f"
)

perc_lucro = st.sidebar.number_input(
    "Lucratividade Líquida Desejada (%)", 
    min_value=0.0, 
    max_value=100.0, 
    value=8.0, 
    step=0.1,
    format="%.2f"
)

# 3. COMPUTAÇÃO E SIMULAÇÃO DINÂMICA
if arquivo_carregado is not None:
    if not os.path.exists(ARQUIVO_TEMPLATE):
        st.error(f"Erro: O arquivo modelo '{ARQUIVO_TEMPLATE}' não foi encontrado no servidor.")
    else:
        try:
            r_adj = perc_receita / 100.0
            l_adj = perc_lucro / 100.0

            # 1. Carregar os dados históricos enviados pelo usuário
            arquivo_carregado.seek(0)
            wb_usuario = openpyxl.load_workbook(io.BytesIO(arquivo_carregado.read()), data_only=True)
            aba_usuario_nome = wb_usuario.sheetnames[0]
            sheet_usuario = wb_usuario[aba_usuario_nome]

            # 2. Carregar o arquivo de template (com fórmulas completas de FC_PROJETADO)
            wb_template = openpyxl.load_workbook(ARQUIVO_TEMPLATE, data_only=False)
            sheet_bal_template = wb_template['BAL_25']
            sheet_fc_template = wb_template['FC_PROJETADO']

            # Copiar os dados históricos do usuário para a aba BAL_25 do template
            for r in range(1, sheet_usuario.max_row + 1):
                for c in range(1, sheet_usuario.max_column + 1):
                    val = sheet_usuario.cell(row=r, column=c).value
                    if val is not None:
                        sheet_bal_template.cell(row=r, column=c, value=val)

            # Injetar os parâmetros configurados pelo utilizador no Excel físico
            sheet_fc_template['B2'] = r_adj
            sheet_fc_template['B3'] = l_adj
            sheet_fc_template['A1'] = f"CLIENTE: {nome_cliente.upper()}"

            # 3. MOTOR DE CÁLCULO DINÂMICO (Para exibir valores calculados em tempo real na tela)
            colunas_meses = [4, 7, 10, 13, 16, 19, 22, 25, 28, 31, 34, 37] # D, G, J, M, P, S, V, Y, AB, AE, AH, AK
            nomes_meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

            # Estrutura para extrair e recalcular em cascata
            dados_recalculados = {}
            
            # Primeiro, vamos ler todos os valores originais do BAL_25 para a memória
            for r in range(4, 106):
                conta = sheet_bal_template.cell(row=r, column=3).value
                tipo = sheet_bal_template.cell(row=r, column=2).value
                
                if conta is not None:
                    valores = []
                    for col in colunas_meses:
                        val = sheet_bal_template.cell(row=r, column=col).value
                        valores.append(float(val) if isinstance(val, (int, float)) else 0.0)
                    
                    dados_recalculados[r] = {
                        "conta": str(conta).strip(),
                        "tipo": str(tipo).strip() if tipo else "",
                        "originais": valores,
                        "projetados": [0.0] * 12
                    }

            # --- PROCESSAMENTO DOS VALORES HISTÓRICOS ---
            receita_total_hist = sum(dados_recalculados[6]["originais"]) # RECEITAS OPERACIONAIS (Linha 6)
            despesa_total_hist = sum(dados_recalculados[98]["originais"]) # DESPESA TOTAL (Linha 98)
            compra_mercadoria_hist = sum(dados_recalculados[14]["originais"]) # Compra de Mercadoria (Linha 14)
            insumos_hist = sum(dados_recalculados[21]["originais"]) # Insumos (Linha 21)

            # --- CÁLCULO DOS INDICADORES E FATOR DE AJUSTE (J4) ---
            receita_projetada = receita_total_hist * (1 + r_adj)
            despesas_variaveis_base = compra_mercadoria_hist + insumos_hist
            
            if despesas_variaveis_base == 0:
                fator_ajuste = 1.0
            else:
                numerador = (receita_projetada * (1 - l_adj)) - (despesa_total_hist - despesas_variaveis_base)
                fator_ajuste = max(0.0, min(1.0, numerador / despesas_variaveis_base))

            # --- REPLICAÇÃO DAS REGRAS DAS FÓRMULAS MÊS A MÊS ---
            for r, info in dados_recalculados.items():
                conta = info["conta"]
                
                # Regra de Receita (Linha 13 - Venda Consumidor Final)
                if r == 13:
                    info["projetados"] = [v * (1 + r_adj) for v in info["originais"]]
                # Regra do Fator de Ajuste (Linhas 17 e 24 - Compra de Mercadoria e Insumos)
                elif r in [17, 24]:
                    info["projetados"] = [v * fator_ajuste for v in info["originais"]]
                # Outras despesas e saldos fixos copiam direto
                else:
                    info["projetados"] = info["originais"].copy()

            # --- RECALCULAR TODAS AS SOMAS/SUBTOTAIS DINAMICAMENTE PARA A TELA ---
            # Vamos aplicar as exatas equações matemáticas das fórmulas do Excel de referência:
            # Receita total operada
            dados_recalculados[11]["projetados"] = dados_recalculados[13]["projetados"].copy() # E11 = E13
            # Custos Operacionais (E15 = E17 a E24)
            for m in range(12):
                dados_recalculados[15]["projetados"][m] = sum(dados_recalculados[r]["projetados"][m] for r in range(17, 25))
            # Faturamento Líquido (E26 = E11 - E15)
            for m in range(12):
                dados_recalculados[26]["projetados"][m] = dados_recalculados[11]["projetados"][m] - dados_recalculados[15]["projetados"][m]
            # Despesas Operacionais (E28 = E29 a E39)
            for m in range(12):
                dados_recalculados[28]["projetados"][m] = sum(dados_recalculados[r]["projetados"][m] for r in range(29, 40))
            # Margem de Contribuição (E41 = E26 - E28)
            for m in range(12):
                dados_recalculados[41]["projetados"][m] = dados_recalculados[26]["projetados"][m] - dados_recalculados[28]["projetados"][m]
            # Despesas de Pessoal (E45 = E46 a E59)
            for m in range(12):
                dados_recalculados[45]["projetados"][m] = sum(dados_recalculados[r]["projetados"][m] for r in range(46, 60))
            # Despesas Fixas Gerais (E60 = E61 a E69)
            for m in range(12):
                dados_recalculados[60]["projetados"][m] = sum(dados_recalculados[r]["projetados"][m] for r in range(61, 70))
            # Despesas Fixas Totais (E43 = E45 + E60)
            for m in range(12):
                dados_recalculados[43]["projetados"][m] = dados_recalculados[45]["projetados"][m] + dados_recalculados[60]["projetados"][m]
            # Lucro Operacional Bruto (E71 = E41 - E43)
            for m in range(12):
                dados_recalculados[71]["projetados"][m] = dados_recalculados[41]["projetados"][m] - dados_recalculados[43]["projetados"][m]
            # Investimentos (E73 = E75 a E77)
            for m in range(12):
                dados_recalculados[73]["projetados"][m] = sum(dados_recalculados[r]["projetados"][m] for r in range(75, 78))
            # Lucro Operacional Líquido (E79 = E71 - E73)
            for m in range(12):
                dados_recalculados[79]["projetados"][m] = dados_recalculados[71]["projetados"][m] - dados_recalculados[73]["projetados"][m]
            # Receitas Não Operacionais (E81 = E83 a E86) -> Com ajuste de receita
            for m in range(12):
                dados_recalculados[81]["projetados"][m] = sum(dados_recalculados[r]["projetados"][m] * (1 + r_adj) for r in range(83, 87))
            # Despesas Não Operacionais (E88 = E90 a E93)
            for m in range(12):
                dados_recalculados[88]["projetados"][m] = sum(dados_recalculados[r]["projetados"][m] for r in range(90, 94))
            # Resultado Líquido Final (E95 = E79 + E81 - E88)
            for m in range(12):
                dados_recalculados[95]["projetados"][m] = dados_recalculados[79]["projetados"][m] + dados_recalculados[81]["projetados"][m] - dados_recalculados[88]["projetados"][m]

            # Receita Total (E99 = E11 + E81)
            for m in range(12):
                dados_recalculados[99]["projetados"][m] = dados_recalculados[11]["projetados"][m] + dados_recalculados[81]["projetados"][m]
            # Despesa Total (E101 = E15 + E28 + E43 + E73 + E88)
            for m in range(12):
                dados_recalculados[101]["projetados"][m] = (
                    dados_recalculados[15]["projetados"][m] +
                    dados_recalculados[28]["projetados"][m] +
                    dados_recalculados[43]["projetados"][m] +
                    dados_recalculados[73]["projetados"][m] +
                    dados_recalculados[88]["projetados"][m]
                )
            # Resultado Líquido Final da aba de baixo (E103 = E99 - E101)
            for m in range(12):
                dados_recalculados[103]["projetados"][m] = dados_recalculados[99]["projetados"][m] - dados_recalculados[101]["projetados"][m]

            # --- VARIÁVEIS DA LINHA 4 CALCULADAS EXATAMENTE DO RECALCULO ---
            rec_proj_final = sum(dados_recalculados[99]["projetados"])
            desp_proj_final = sum(dados_recalculados[101]["projetados"])
            res_proj_final = rec_proj_final - desp_proj_final
            margem_proj_final = (res_proj_final / rec_proj_final) if rec_proj_final != 0 else 0.0

            # --- APRESENTAÇÃO DOS INDICADORES NA TELA ---
            st.markdown("### 📊 Indicadores Financeiros do Cenário (Equivalente à Linha 4)")
            col_r1, col_r2, col_r3, col_r4, col_r5 = st.columns(5)
            with col_r1:
                st.metric("Receita Projetada (B4)", f"R$ {rec_proj_final:,.2f}")
            with col_r2:
                st.metric("Despesa Projetada (D4)", f"R$ {desp_proj_final:,.2f}")
            with col_r3:
                st.metric("Resultado Projetado (F4)", f"R$ {res_proj_final:,.2f}")
            with col_r4:
                st.metric("Margem Projetada (H4)", f"{margem_proj_final * 100:.2f}%")
            with col_r5:
                st.metric("Fator de Ajuste (J4)", f"{fator_ajuste * 100:.2f}%")

            # --- CONSTRUÇÃO DA TABELA DINÂMICA INTERATIVA COMPLETA ---
            st.markdown("### 📈 Visualização Dinâmica da Planilha FC_PROJETADO")
            
            linhas_visualizacao = []
            for r in sorted(dados_recalculados.keys()):
                info = dados_recalculados[r]
                proj_vals = info["projetados"]
                total_projetado = sum(proj_vals)
                
                linhas_visualizacao.append({
                    "Conta / Hierarquia": info["conta"],
                    "Tipo": info["tipo"],
                    "Total Projetado": total_projetado,
                    **{nomes_meses[i]: proj_vals[i] for i in range(12)}
                })
                
            df_visual = pd.DataFrame(linhas_visualizacao)
            
            # Formatação de visualização de moeda em tela
            df_formatado = df_visual.copy()
            for col in ["Total Projetado"] + nomes_meses:
                df_formatado[col] = df_formatado[col].apply(lambda x: f"R$ {x:,.2f}" if isinstance(x, (int, float)) else x)
                
            st.dataframe(df_formatado, use_container_width=True, height=450)

            # --- EXCEL FÍSICO COM FÓRMULAS ---
            # Modifica as células de receita e custo dinamicamente no arquivo para exportação
            for r in range(1, sheet_fc_template.max_row + 1):
                cell_conta = sheet_bal_template.cell(row=r, column=3).value
                if cell_conta == "1.1 Venda Consumidor Final":
                    for col_idx in colunas_meses:
                        val_hist = sheet_bal_template.cell(row=r, column=col_idx).value
                        if isinstance(val_hist, (int, float)):
                            sheet_fc_template.cell(row=r, column=col_idx, value=val_hist * (1 + r_adj))
                elif cell_conta in ["3.1 Compra de Mercadoria", "3.8 Insumos"]:
                    for col_idx in colunas_meses:
                        val_hist = sheet_bal_template.cell(row=r, column=col_idx).value
                        if isinstance(val_hist, (int, float)):
                            sheet_fc_template.cell(row=r, column=col_idx, value=f"={sheet_bal_template.title}!{openpyxl.utils.get_column_letter(col_idx)}{r}*$J$4")

            # Salvar em formato binário para download
            output_web = io.BytesIO()
            wb_template.save(output_web)
            output_web.seek(0)

            st.markdown("---")
            st.markdown("### 📥 Exportar Resultados")
            st.download_button(
                label="📥 Exportar Planilha com FC_PROJETADO Oficial (Com Fórmulas)",
                data=output_web,
                file_name=f"FC_PROJETADO_{nome_cliente.upper()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Erro ao processar as fórmulas com o modelo: {e}")
else:
    st.info("ℹ️ Insira os dados de simulação e carregue a planilha base para ver os resultados.")
